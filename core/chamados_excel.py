import logging
import json
import re
import unicodedata
from datetime import datetime
from pathlib import Path

from django.conf import settings
from django.utils import timezone
from openpyxl import load_workbook

from .models import ERPUser, Ticket

logger = logging.getLogger(__name__)

MONTH_TOKENS = {
    1: ('jan', 'janeiro', '01'),
    2: ('fev', 'fevereiro', '02'),
    3: ('mar', 'marco', '03'),
    4: ('abr', 'abril', '04'),
    5: ('mai', 'maio', '05'),
    6: ('jun', 'junho', '06'),
    7: ('jul', 'julho', '07'),
    8: ('ago', 'agosto', '08'),
    9: ('set', 'setembro', '09'),
    10: ('out', 'outubro', '10'),
    11: ('nov', 'novembro', '11'),
    12: ('dez', 'dezembro', '12'),
}


def _normalize(value: str) -> str:
    raw = (value or '').strip().lower()
    base = unicodedata.normalize('NFKD', raw).encode('ascii', 'ignore').decode('ascii')
    base = re.sub(r'\s+', ' ', base)
    return base


def _build_header_map(cells: list[str]) -> dict[str, int]:
    wanted = {
        'ti': None,
        'data': None,
        'contato': None,
        'setor': None,
        'notificacao': None,
        'prioridade': None,
        'falha': None,
        'acao': None,
        'fechado': None,
        'tempo': None,
        'acao_eficaz': None,
    }
    for idx, raw in enumerate(cells, start=1):
        key = _normalize(str(raw or ''))
        if key == 'ti':
            wanted['ti'] = idx
        elif key == 'data':
            wanted['data'] = idx
        elif key == 'contato':
            wanted['contato'] = idx
        elif key == 'setor':
            wanted['setor'] = idx
        elif key == 'notificacao':
            wanted['notificacao'] = idx
        elif key == 'prioridade':
            wanted['prioridade'] = idx
        elif key == 'falha':
            wanted['falha'] = idx
        elif key in {'acao / correcao', 'acao/correcao', 'acao correcao'}:
            wanted['acao'] = idx
        elif key == 'fechado':
            wanted['fechado'] = idx
        elif key == 'tempo':
            wanted['tempo'] = idx
        elif key == 'acao eficaz':
            wanted['acao_eficaz'] = idx
    return wanted


def _resolve_sheet(workbook, event_dt: datetime):
    best_sheet = workbook.active
    best_score = -1
    month_tokens = MONTH_TOKENS.get(event_dt.month, ())
    year_text = str(event_dt.year)
    for sheet in workbook.worksheets:
        normalized = _normalize(sheet.title)
        score = 0
        if year_text in normalized:
            score += 3
        if any(token in normalized for token in month_tokens):
            score += 3
        if score > best_score:
            best_score = score
            best_sheet = sheet
    return best_sheet


def _find_header(sheet):
    for row_idx in range(1, 8):
        raw = [sheet.cell(row=row_idx, column=col).value for col in range(1, 30)]
        header_map = _build_header_map(raw)
        if header_map['data'] and header_map['contato'] and header_map['notificacao']:
            return row_idx, header_map

    return 1, {
        'ti': 1,
        'data': 2,
        'contato': 3,
        'setor': 4,
        'notificacao': 5,
        'prioridade': 6,
        'falha': 7,
        'acao': 8,
        'fechado': 9,
        'tempo': 10,
        'acao_eficaz': 11,
    }


def _find_next_row(sheet, header_row: int, header_map: dict[str, int]) -> int:
    key_cols = [header_map[k] for k in ('data', 'contato', 'notificacao', 'fechado') if header_map.get(k)]
    if not key_cols:
        return header_row + 1
    row = header_row + 1
    while True:
        has_value = any((sheet.cell(row=row, column=col).value not in (None, '')) for col in key_cols)
        if not has_value:
            return row
        row += 1


def _format_dt(dt: datetime) -> str:
    return timezone.localtime(dt).strftime('%d/%m/%Y %H:%M')


def _format_duration(opened_at: datetime, closed_at: datetime) -> str:
    seconds = int(max((closed_at - opened_at).total_seconds(), 0))
    minutes = seconds // 60
    hours = minutes // 60
    mins = minutes % 60
    return f'{hours:02d}:{mins:02d}'


def _pending_queue_path() -> Path:
    configured = getattr(settings, 'CHAMADOS_XLSX_PENDING_PATH', '') or ''
    if configured:
        return Path(configured)
    return Path(settings.BASE_DIR) / 'data' / 'chamados_excel_pending.jsonl'


def _read_pending_entries(path: Path) -> list[dict]:
    if not path.exists():
        return []
    items: list[dict] = []
    for line in path.read_text(encoding='utf-8').splitlines():
        raw = (line or '').strip()
        if not raw:
            continue
        try:
            parsed = json.loads(raw)
            if isinstance(parsed, dict):
                items.append(parsed)
        except Exception:
            continue
    return items


def _write_pending_entries(path: Path, entries: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not entries:
        if path.exists():
            path.unlink()
        return
    content = '\n'.join(json.dumps(item, ensure_ascii=False) for item in entries) + '\n'
    path.write_text(content, encoding='utf-8')


def _queue_pending_entry(event_dt: datetime, values: dict, ticket_id: int) -> None:
    path = _pending_queue_path()
    entries = _read_pending_entries(path)
    entries.append(
        {
            'ticket_id': ticket_id,
            'event_dt': timezone.localtime(event_dt).isoformat(),
            'values': values,
        }
    )
    _write_pending_entries(path, entries)
    logger.warning('Lancamento do chamado #%s enfileirado para reprocesso: %s', ticket_id, path)


def _append_values_to_workbook(workbook_path: Path, event_dt: datetime, values: dict) -> None:
    wb = load_workbook(workbook_path)
    sheet = _resolve_sheet(wb, timezone.localtime(event_dt))
    header_row, header_map = _find_header(sheet)
    target_row = _find_next_row(sheet, header_row, header_map)
    for key, col in header_map.items():
        if not col:
            continue
        if key not in values:
            continue
        sheet.cell(row=target_row, column=col, value=values[key])
    wb.save(workbook_path)


def _flush_pending_entries(workbook_path: Path) -> None:
    path = _pending_queue_path()
    entries = _read_pending_entries(path)
    if not entries:
        return

    remaining: list[dict] = []
    for idx, entry in enumerate(entries):
        try:
            event_dt = datetime.fromisoformat(str(entry.get('event_dt', '')))
            values = entry.get('values') or {}
            _append_values_to_workbook(workbook_path, event_dt, values)
        except PermissionError:
            remaining.append(entry)
            remaining.extend(entries[idx + 1 :])
            break
        except Exception:
            logger.exception('Falha ao reprocessar lancamento enfileirado: %s', entry)

    _write_pending_entries(path, remaining)


def append_chamado_event_to_excel(
    *,
    ticket: Ticket,
    opened_at: datetime,
    closed_at: datetime,
    failure_type: str,
    action_text: str,
) -> bool:
    workbook_path = Path(getattr(settings, 'CHAMADOS_XLSX_PATH', '') or '')
    if not workbook_path:
        return False
    if not workbook_path.exists():
        logger.warning('Arquivo de chamados nao encontrado: %s', workbook_path)
        return False

    try:
        _flush_pending_entries(workbook_path)

        creator = ticket.created_by
        username = (creator.username if creator else '') or ''
        erp_user = ERPUser.objects.filter(username__iexact=username).first() if username else None
        contato = (erp_user.full_name if erp_user and erp_user.full_name else username) or '-'
        setor = (erp_user.department if erp_user else '') or ''

        failure_map = {
            Ticket.FailureType.NS: 'N/S',
            Ticket.FailureType.EQUIPAMENTO: 'Equipamento',
            Ticket.FailureType.SOFTWARE: 'Software',
        }
        failure_label = failure_map.get(failure_type, failure_type or '')

        values = {
            'ti': '',
            'data': _format_dt(opened_at),
            'contato': contato,
            'setor': setor,
            'notificacao': ticket.description or '',
            'prioridade': ticket.get_urgency_display(),
            'falha': failure_label,
            'acao': action_text or '',
            'fechado': _format_dt(closed_at),
            'tempo': _format_duration(opened_at, closed_at),
            'acao_eficaz': '',
        }

        _append_values_to_workbook(workbook_path, closed_at, values)
        return True
    except PermissionError:
        _queue_pending_entry(closed_at, values, ticket.id)
        return False
    except Exception:
        logger.exception('Falha ao escrever chamado #%s na planilha', ticket.id)
        return False
